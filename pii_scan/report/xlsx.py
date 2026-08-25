# -*- coding: utf-8 -*-
"""Реестр в XLSX — под перечень обрабатываемых ПДн.

Листы разделены по вердикту, поэтому колонки «вердикт» в них нет. Названия
листов — предикаты поля, чтобы строка читалась целиком: поле «содержит ПДн»
либо «требует проверки».

  Содержит ПДн      — уверенные находки, строка на поле;
  Требует проверки  — пограничные находки, материал для ручного разбора;
  Все поля          — все поля всех таблиц, только при full_inventory;
  Сводка            — по таблицам, счётчики прогона и состав книги.
"""
from __future__ import annotations

from typing import List, Optional

from ..model import Finding, ScanResult, VERDICT_TITLES

REGISTRY_HEADER = [
    "Источник", "СУБД", "База данных", "Таблица", "Поле", "Тип поля",
    "Категория ПДн", "Вид ПДн", "Третьи лица", "Уверенность", "Основание",
    "Совпало значений", "Строк в таблице (оценка)", "Комментарий к полю",
]
# Колонка примеров добавляется только когда их попросили ключом --examples:
# пустой столбец в реестре читается как «примеров не нашлось», а не как
# «их сознательно не выгружали».
EXAMPLES_COLUMN = "Примеры (маскированные)"

INVENTORY_HEADER = [
    "Источник", "СУБД", "База данных", "Таблица", "Поле", "Тип поля",
    "Вердикт", "Категория ПДн", "Вид ПДн", "Уверенность", "Основание",
    "Совпало значений", "Комментарий к полю",
]

SUMMARY_HEADER = [
    "Источник", "База данных", "Таблица", "Полей с ПДн", "Категории",
    "Спецкатегории", "Третьи лица", "Строк в таблице (оценка)", "Уверенность",
]


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "для выгрузки XLSX нужен openpyxl (pip install openpyxl)"
        ) from exc
    return __import__("openpyxl")


def _row(finding: Finding, source_type: str,
         with_examples: bool = False) -> List:
    examples: List[str] = []
    for code in finding.codes:
        for ex in finding.hits[code].examples:
            if ex not in examples:
                examples.append(ex)
    return [
        finding.ref.source,
        source_type,
        finding.ref.database,
        finding.ref.table,
        finding.ref.full_column,
        finding.ref.data_type,
        ", ".join(finding.categories) or "—",
        ", ".join(finding.titles),
        "да" if finding.third_party else "",
        round(finding.score, 2),
        finding.basis,
        finding.coverage,
        finding.rows_total if finding.rows_total is not None else "",
        finding.ref.comment,
    ] + (["; ".join(examples[:3])] if with_examples else [])


def write_xlsx(result: ScanResult, path: str) -> None:
    examples = bool(result.options.get("examples_per_hit"))
    registry_header = REGISTRY_HEADER + ([EXAMPLES_COLUMN] if examples else [])
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    types = {s["name"]: s["type"] for s in result.sources}

    wb = openpyxl.Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="44546A")
    special_fill = PatternFill("solid", fgColor="FCE4D6")

    def setup(ws, header: List[str]) -> None:
        ws.append(header)
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"

    def autosize(ws, widths: Optional[List[int]] = None) -> None:
        for idx in range(1, ws.max_column + 1):
            letter = get_column_letter(idx)
            if widths and idx <= len(widths):
                ws.column_dimensions[letter].width = widths[idx - 1]
                continue
            longest = max(
                (len(str(c.value)) for c in ws[letter] if c.value is not None),
                default=10,
            )
            ws.column_dimensions[letter].width = min(max(longest + 2, 10), 45)

    # --- Содержит ПДн ---
    # Лист вердикта «ПДн»: сюда попадает то, что идёт в перечень
    # обрабатываемых персональных данных без дополнительного разбора.
    ws = wb.active
    ws.title = "Содержит ПДн"
    setup(ws, registry_header)
    for table in result.pii_tables:
        for finding in table.pii_findings:
            ws.append(_row(finding, types.get(finding.ref.source, ""), examples))
            if "специальные" in finding.categories:
                for cell in ws[ws.max_row]:
                    cell.fill = special_fill
    autosize(ws)

    # --- Требует проверки ---
    # Сначала те, где сработало содержимое: их разбирают глазами. Ниже —
    # держащиеся на одном имени поля: там смотреть не на что, вопрос к
    # разработчикам. Различить их можно и по колонке «Основание», но
    # вперемешку список читается как одна общая куча работы.
    ws = wb.create_sheet("Требует проверки")
    setup(ws, registry_header)
    for finding in result.pending_findings:
        ws.append(_row(finding, types.get(finding.ref.source, ""), examples))
    autosize(ws)

    # --- Все поля (только в режиме полной описи) ---
    if result.options.get("full_inventory"):
        ws = wb.create_sheet("Все поля")
        setup(ws, INVENTORY_HEADER)
        clean_fill = PatternFill("solid", fgColor="F2F2F2")
        for table in result.tables:
            for finding in sorted(table.findings,
                                  key=lambda f: (-f.score, f.ref.full_column)):
                ws.append([
                    finding.ref.source,
                    types.get(finding.ref.source, ""),
                    finding.ref.database,
                    finding.ref.table,
                    finding.ref.full_column,
                    finding.ref.data_type,
                    VERDICT_TITLES[finding.verdict],
                    ", ".join(finding.categories) or "—",
                    finding.summary_kind,
                    round(finding.score, 2) if finding.score else "",
                    finding.basis,
                    finding.coverage,
                    finding.ref.comment,
                ])
                if finding.verdict == "no":
                    for cell in ws[ws.max_row]:
                        cell.fill = clean_fill
                elif "специальные" in finding.categories:
                    for cell in ws[ws.max_row]:
                        cell.fill = special_fill
        autosize(ws)

    # --- Сводка ---
    ws = wb.create_sheet("Сводка")
    setup(ws, SUMMARY_HEADER)
    for table in result.pii_tables:
        ws.append([
            table.source, table.database, table.table,
            len(table.pii_findings), ", ".join(table.categories) or "—",
            "да" if table.has_special else "",
            "да" if table.third_party else "",
            table.rows_total if table.rows_total is not None else "",
            round(table.score, 2),
        ])
        if table.has_special:
            for cell in ws[ws.max_row]:
                cell.fill = special_fill
    ws.append([])
    ws.append(["Обследование проведено", result.started_at])
    ws.append(["Длительность, с", result.duration_sec])
    ws.append(["Режим", "инвентаризация схемы"
               if result.options.get("dry_run") else "выборка значений"])
    ws.append(["Таблиц с ПДн", len(result.pii_tables)])
    ws.append(["Полей с ПДн", sum(len(t.pii_findings) for t in result.tables)])
    ws.append(["Таблиц на проверку", len(result.maybe_tables)])
    ws.append(["Полей на ручной разбор", len(result.pending_findings)])
    ws.append(["  подтверждено значениями", len(result.pending_confirmed)])
    ws.append(["  только по имени поля", len(result.pending_by_name)])

    # Какой лист что означает. Названия листов сами по себе не говорят,
    # где окончательный вывод, а где материал для разбора.
    ws.append([])
    ws.append(["Листы книги", "Вердикт", "Что внутри"])
    for cell in ws[ws.max_row]:
        cell.font = head_font
    for row in [
        ["Содержит ПДн", "ПДн",
         "уверенные находки, идут в перечень обрабатываемых ПДн"],
        ["Требует проверки", "требует проверки",
         "сигнала недостаточно для вывода, нужен человек"],
        ["Все поля", "все три",
         "полная опись, включая чистые поля (ключ --full-inventory)"
         if result.options.get("full_inventory")
         else "лист не создан: нужен ключ --full-inventory"],
    ]:
        ws.append(row)
    autosize(ws)

    wb.save(path)
